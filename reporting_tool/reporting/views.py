from django.conf import settings
from django.urls import reverse, reverse_lazy
from django.shortcuts import redirect, render, get_object_or_404
from django.http import Http404, HttpResponse, HttpResponseServerError, JsonResponse
from django.template import loader
from django.template.loader import render_to_string
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.views import PasswordChangeDoneView
from django.contrib import messages
from .models import Finding, DB_CWE, DB_OWASP, Customer, Report, Finding_Template, UserProfile, Appendix, Deliverable
from martor.utils import LazyEncoder

from .forms import Add_findings, AddOWASP, AddCustomer, AddReport, OWASP_Questions, NewFindingTemplateForm, UserForm, UploadNmapForm, UploadOpenVASForm, AppendixForm


from dotenv import load_dotenv
from collections import Counter
from time import sleep
from openai import OpenAI
import xml.etree.ElementTree as ET
import datetime
import pypandoc
import os
import json
import base64
import pathlib
import textwrap
import uuid
import logging

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

from config.configs import USER_CONFIG, DJANGO_CONFIG, DELIVERABLES_CONFIG

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------
# https://github.com/agusmakmun/django-markdown-editor/wiki
# ----------------------------------------------------------------------

@login_required
def markdown_uploader(request):
    """
    Makdown image upload for locale storage
    and represent as json to markdown editor.
    """
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if 'markdown-image-upload' in request.FILES:
            image = request.FILES['markdown-image-upload']
            image_types = [
                'image/png', 'image/jpg',
                'image/jpeg', 'image/pjpeg', 'image/gif'
            ]
            if image.content_type not in image_types:
                data = json.dumps({
                    'status': 405,
                    'error': ('Bad image format.')
                }, cls=LazyEncoder)
                return HttpResponse(
                    data, content_type='application/json', status=405)

            if image.size > settings.MAX_IMAGE_UPLOAD_SIZE:
                to_MB = settings.MAX_IMAGE_UPLOAD_SIZE / (1024 * 1024)
                data = json.dumps({
                    'status': 405,
                    'error': ('Maximum image file is %(size)s MB.') % {'size': to_MB}
                }, cls=LazyEncoder)
                return HttpResponse(
                    data, content_type='application/json', status=405)
            
            image_content_base64 = base64.b64encode(image.read()).decode('utf-8')
             
            image_content_base64_final = 'data:' + image.content_type +';base64,' + image_content_base64

            data = json.dumps({
                'status': 200,
                'link': image_content_base64_final,
                'name': image.name
                })

            # img_uuid = "{0}-{1}".format(uuid.uuid4().hex[:10], image.name.replace(' ', '-'))
            # tmp_file = os.path.join(settings.MARTOR_UPLOAD_PATH, img_uuid)
            # def_path = default_storage.save(tmp_file, ContentFile(image.read()))
            # img_url = os.path.join(settings.MEDIA_URL, def_path)

            # data = json.dumps({
            #     'status': 200,
            #     'link': img_url,
            #     'name': image.name
            # })
            return HttpResponse(data, content_type='application/json')
        return HttpResponse(('Invalid request!'))
    return HttpResponse(('Invalid request!'))
#---------------------------------------------------
#                   CHATGPT
# --------------------------------------------------

load_dotenv()
client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
)

def format_chatgpt_output(ai_response, form):
    ai_response = ai_response.replace("'", '"')
    report_data = json.loads(ai_response)
    if type(report_data['References']) is not str:
        for index, reference in enumerate(report_data['References']):
            if reference[0] != "-":
                report_data['References'][index] = "- " + reference
        report_data['References'] = "\n".join(report_data['References'])
        
    if type(report_data['Recommendation']) is not str:
        for index, recommendation in enumerate(report_data['Recommendation']):
            if recommendation[0] != "-":
                report_data['Recommendation'][index] = "- " + recommendation
        
        report_data['Recommendation'] = "\n".join(report_data['Recommendation'])
    

    report_data['owasp'] = form.data["owasp"]
    report_data['location'] = form.data["affected_url"]
    return report_data

#---------------------------------------------------
#                   AUTH
# --------------------------------------------------
class CustomPasswordChangeView(PasswordChangeView):
    template_name = 'registration/password_change_form.html'  # Your custom template
    success_url = reverse_lazy('custom_password_change_done')  # URL to redirect after a successful password change
    
class CustomPasswordChangeDoneView(PasswordChangeDoneView):
    # Specify your custom template (optional)
    template_name = 'registration/password_change_done.html'
    
    def dispatch(self, *args, **kwargs):
        # Example: Add a success message using Django's messages framework
        messages.success(self.request, 'ANJING ANJING ANJINGS')
        return redirect(home)
    

#---------------------------------------------------
#                   HOME
# --------------------------------------------------
@login_required
def home(request):
    template = loader.get_template('index.html')
    return HttpResponse(template.render())

#---------------------------------------------------
#                   USER
# --------------------------------------------------
@login_required
def user_view(request):
    return render(request, 'user/user_view.html', {
        'user_profile': request.user
    })

@login_required
def user_edit(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=request.user)
        
        if user_form.is_valid() :
            user_form.save()
            return redirect('user_view')  # Redirect to a success page or the profile page
    else:
        user_form = UserForm(instance=request.user)
    
    context = {
        'user_form': user_form,
    }
    return render(request, 'user/user_edit.html', context)

#---------------------------------------------------
#                   SETTINGS
# --------------------------------------------------
@login_required
def update_preference(request):
    if request.method == "POST" and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:            
            new_font_size = request.POST.get('fontSize', '')
            new_background_colour = request.POST.get('backgroundColour', '')
            new_font_colour = request.POST.get('fontColour', '')
            new_font_type = request.POST.get('fontType', '')
            new_character_spacing = request.POST.get('characterSpacing', 0)
            new_line_height = request.POST.get('lineHeight', 1.5)
            profile = request.user.userprofile
            profile.font_color = new_font_colour
            profile.background_color = new_background_colour
            profile.font_size = new_font_size
            profile.font_type = new_font_type
            profile.character_spacing = new_character_spacing
            profile.line_height = new_line_height
            profile.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'success': False}, status=400)

#---------------------------------------------------
#                   FINDINGS
# --------------------------------------------------
@login_required
def findings_open_list(request):
    findings = Finding.objects.filter(status="Open")
    template = loader.get_template('findings/findings_display.html')
    context = {
        'findings':findings,
        'status': "Open"
    }
    return HttpResponse(template.render(context, request))

@login_required
def findings_closed_list(request):
    findings = Finding.objects.filter(status="Closed")
    template = loader.get_template('findings/findings_display.html')
    context = {
        'findings':findings,
        'status': "Closed"
    }
    return HttpResponse(template.render(context, request))

@login_required
def add_finding(request,pk):
    Report_query = get_object_or_404(Report, pk=pk)

    if request.method == 'POST':
        form = Add_findings(request.POST)
        
        if form.is_valid():
            finding = form.save(commit=False) 
            finding.report = Report_query       
            finding.finding_id = uuid.uuid4()
            finding.save()
            
        return redirect('report_finding', pk=pk)
        
    else:
        form = Add_findings()
        form.fields['description'].initial = "TBC"
        # form.fields['impact'].initial = "TBC"
        form.fields['recommendation'].initial = "TBC"
        form.fields['references'].initial = "TBC"
        form.fields['location'].initial = "TBC"
        form.fields['status'].initial = "Open"
        form.fields['poc'].initial = "TBC"
        template = 'findings/findings_add.html'
        context = {
            'form': form,
            'Report_query': Report_query
        }

    return render(request, template, context)


@login_required
def add_finding_from_gpt(request,pk):
    report_data_json = request.session.get('report_data', None)
    # Check if report_data_json is in the session
    if report_data_json is not None:
        # Deserialize report_data_json from JSON to a dictionary
        report_data_dict = json.loads(report_data_json)
    else:
        # Handle the case where report_data is not in the session
        report_data_dict = {}
    Report_query = get_object_or_404(Report, pk=pk)

    if request.method == 'POST':
        form = Add_findings(request.POST)
        
        if form.is_valid():
            finding = form.save(commit=False) 
            finding.report = Report_query       
            finding.finding_id = uuid.uuid4()
            finding.save()
            
        return redirect('report_finding', pk=pk)
        
    else:
        form = Add_findings()
        form.fields['description'].initial = report_data_dict['Description']
        # form.fields['impact'].initial = report_data_dict['Impact']
        form.fields['recommendation'].initial = report_data_dict['Recommendation']
        form.fields['references'].initial = report_data_dict['References']
        form.fields['owasp'].initial = report_data_dict['owasp']
        form.fields['cvss_score'].initial = report_data_dict['CVSS Score Number']
        form.fields['severity'].initial = report_data_dict['Criticality']
        form.fields['location'].initial = report_data_dict['location']
        form.fields['status'].initial = "Open"
        form.fields['poc'].initial = "TBC"
        template = 'findings/findings_add.html'
        context = {
            'form': form,
            'Report_query': Report_query
        }

    return render(request, template, context)

@login_required
def edit_finding(request, pk):
    
    finding = get_object_or_404(Finding, pk=pk)
    print(finding.severity)
    Report_query = get_object_or_404(Report, pk=finding.report.id)
    
    if request.method == 'POST':
        form = Add_findings(request.POST, instance=finding)
        
        if form.is_valid():
            finding = form.save(commit=False)            
            finding.finding_id = uuid.uuid4()
            finding.save()
            
        return redirect('/findings/view/'+ str(pk))
        
    else:
        form = Add_findings(instance=finding)
        template = loader.get_template('findings/findings_add.html')
        context = {
            'form': form,
            'Report_query': Report_query
        }

    return render(request, 'findings/findings_add.html', context)

@login_required
def view_finding(request, pk):
    finding = get_object_or_404(Finding, pk=pk)
    template = loader.get_template('findings/findings_view.html')
    context = {
        'finding':finding
    }
    return HttpResponse(template.render(context, request))

@login_required
def finding_delete(request,pk):
    Finding.objects.filter(pk=pk).delete()
    next_url = request.GET.get('next', '/')
    return redirect(next_url)

@login_required
def findingtotemplate(request,pk,reportpk):
    finding = get_object_or_404(Finding, pk=pk)
    template_title = finding.title
    template_severity = finding.severity
    template_cvss_vector = finding.cvss_vector
    template_cvss_score = finding.cvss_score
    template_description = finding.description
    # template_impact = finding.impact
    template_recommendation = finding.recommendation
    template_references = finding.references
    template_owasp = finding.owasp
    
    template_to_DB = Finding_Template(
        title = template_title,
        severity = template_severity,
        cvss_vector = template_cvss_vector,
        cvss_score = template_cvss_score,
        description = template_description,
        # impact = template_impact,
        recommendation = template_recommendation,
        references = template_references,
        owasp = template_owasp,
    )
    
    template_to_DB.save()
    
    # In your view
    request.session['success'] = True
    
    return redirect(reverse('report_finding', kwargs={'pk': reportpk}))
    

@login_required
def initial_add_finding(request, pk):

    if request.method == 'POST':
        form = OWASP_Questions(request.POST)        
        if form.is_valid():
            DB_owasp_query = get_object_or_404(DB_OWASP, pk=int(form.data["owasp"]))
            info = {}
            prompt = """
You are a professional penetration tester. You have performed a penetration test on a particular company. You will be provided details on a particular vulnerability that was found during the penetration test. 
Please provide content for a detailed finding report based on the information that will be given. The report will be included as part of the complete penetration testing report.
There are 6 sections to be included. 
- Description (explain how it works and add the impact of the vulnerability in here)
- Recommendation (in bullet point form)
- CVSS Score Number (assign a CVSS score based on your understanding of the vulnerability. Only include the number without any other explanation)
- Criticality (critical, high, medium, low, info only)
- References (this should include full links for reference materials.). 

Ignore all HTML Tags. Output in JSON Format with each Section header, eg Description, Recommendation as keys. Do not include any newline characters in the response.
            """
            str_info = ""
            for response in form:
                if response.field.label == "OWASP":
                    info[response.field.label] = "%s : %s" %(DB_owasp_query.owasp_full_id , DB_owasp_query.owasp_name)
                else:
                    info[response.field.label] = response.data
            for label, value in info.items():
                temp = "{label}: {value}\n"
                str_info = str_info + temp.format(label = label, value = value)
            
            while True:
                try:
                    completion = client.chat.completions.create(
                        model="gpt-4",
                        # model="gpt-3.5-turbo",
                        messages=[
                            {"role": "system", "content": prompt},
                            {"role": "user", "content": str_info}
                        ]
                    )
                    ai_response = completion.choices[0].message.content
                    
                    report_data = format_chatgpt_output(ai_response,form)
                    break
                except json.JSONDecodeError:
                    sleep(0.5)
            
            
            report_data_json = json.dumps(report_data)
            
            request.session['report_data'] = report_data_json
            redirect_url = f'/findings/add_gpt/{pk}'
            
        return redirect(redirect_url)
        
    else:
        form = OWASP_Questions()
        template = 'findings/findings_initital_add.html'
        context = {
            'form': form,
        }

    return render(request, template, context)
    
#---------------------------------------------------
#                   TEMPLATES
# --------------------------------------------------

@login_required
def template_add(request):
    if request.method == 'POST':
        form = NewFindingTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.save()

            if '_finish' in request.POST:
                return redirect('template_list')
            elif '_next' in request.POST:
                return redirect('template_add')
        
    else:
        form = NewFindingTemplateForm()
        form.fields['description'].initial = "TBC"
        # form.fields['impact'].initial = "TBC"
        form.fields['recommendation'].initial = "TBC"
        form.fields['references'].initial = "TBC"
        form.fields['owasp'].initial = '1'

    return render(request, 'findings/template_add.html', {
        'form': form
    })


@login_required
def template_edit(request, pk):
    
    template = get_object_or_404(Finding_Template, pk=pk)
    
    if request.method == 'POST':
        form = NewFindingTemplateForm(request.POST, instance=template)
        if form.is_valid():
            template = form.save(commit=False)
            template.save()

            if '_finish' in request.POST:
                return redirect('template_list')
            elif '_next' in request.POST:
                return redirect('template_add')
        
    else:
        form = NewFindingTemplateForm(instance=template)

    return render(request, 'findings/template_add.html', {
        'form': form
    })
    

@login_required
def template_duplicate(request,pk):
    template_query = get_object_or_404(Finding_Template, pk=pk)
    
    if request.method == 'POST':
        form = NewFindingTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.save()

            if '_finish' in request.POST:
                return redirect('template_list')
            elif '_next' in request.POST:
                return redirect('template_add')
        
    else:
        form = NewFindingTemplateForm()
        form.fields['title'].initial = template_query.title
        form.fields['severity'].initial = template_query.severity
        form.fields['cvss_score'].initial = template_query.cvss_score
        form.fields['cvss_vector'].initial = template_query.cvss_vector
        form.fields['description'].initial = template_query.description
        # form.fields['impact'].initial = template_query.impact
        form.fields['recommendation'].initial = template_query.recommendation
        form.fields['references'].initial = template_query.references
        form.fields['owasp'].initial = template_query.owasp

    return render(request, 'findings/template_add.html', {
        'form': form
    })
    

@login_required
def template_delete(request, pk):
    template = get_object_or_404(Finding_Template, pk=pk)
    template.delete()
    return redirect('template_list')

@login_required
def template_list(request):
    Templates = Finding_Template.objects.order_by('title')
    
    return render(request, 'findings/template_list.html', {
        'Templates': Templates
    })

@login_required
def template_view(request, pk):
    finding = get_object_or_404(Finding_Template, pk=pk)
    template = 'findings/template_view.html'
    context = {
        'finding':finding
    }
    return render(request, template, context)

@login_required
def templateaddfinding(request,pk):

    DB_report_query = get_object_or_404(Report, pk=pk)
    DB_findings_query = Finding_Template.objects.order_by('title')

    return render(request, 'findings/templateaddfinding.html', {'DB_findings_query': DB_findings_query, 'DB_report_query': DB_report_query})



@login_required
def templateaddreport(request,pk,reportpk):

    DB_report_query = get_object_or_404(Report, pk=reportpk)
    DB_finding_template_query = get_object_or_404(Finding_Template, pk=pk)

    # save template in DB
    finding_uuid = uuid.uuid4()
    finding_status = "Open"
    finding_to_DB = Finding(report=DB_report_query, finding_id=finding_uuid, title=DB_finding_template_query.title, severity=DB_finding_template_query.severity, cvss_vector=DB_finding_template_query.cvss_vector, cvss_score=DB_finding_template_query.cvss_score, description=DB_finding_template_query.description, status=finding_status, recommendation=DB_finding_template_query.recommendation, references=DB_finding_template_query.references, owasp=DB_finding_template_query.owasp)

    finding_to_DB.save()

    return redirect('report_finding', pk=reportpk)


@login_required
def template_findings_autocomplete(request):
    if 'term' in request.GET:
        qs = Finding_Template.objects.filter(title__icontains=request.GET.get('term'))
        findings = [{
            'label': finding.title,  
            'value': finding.title,  
            'severity': finding.severity,  
            'cvss_vector': finding.cvss_vector,
            'cvss_score': finding.cvss_score,
            'description': finding.description,
            # 'impact': finding.impact,
            'recommendation': finding.recommendation,
            'references': finding.references,
            'owasp': finding.owasp.owasp_id,
            
        } for finding in qs]
        return JsonResponse(findings, safe=False)
    return JsonResponse([], safe=False)


#---------------------------------------------------
#                   OWASPS
# --------------------------------------------------

@login_required
def owasp_list(request):
    owasps = DB_OWASP.objects.all().values()
    template = loader.get_template('owasp/owasp_list.html')
    context = {
        'owasps':owasps
    }
    return HttpResponse(template.render(context, request))

@login_required
def owasp_add(request):
    if request.method == 'POST':
        form = AddOWASP(request.POST)
        
        if form.is_valid():
            owasp = form.save(commit=False)            
            owasp.save()
            
        return redirect('/owasp/list')
        
    else:
        form = AddOWASP()
        template = 'owasp/owasp_add.html'
        context = {
            'form': form
        }

    return render(request, template , context)

@login_required
def owasp_edit(request,pk):
    
    owasp = get_object_or_404(DB_OWASP, pk=pk)
    
    if request.method == 'POST':
        form = AddOWASP(request.POST, instance=owasp)
        
        if form.is_valid():
            owasp = form.save(commit=False)            
            owasp.save()
            
        return redirect('/owasp/list')
        
    else:
        form = AddOWASP(instance=owasp)
        template = 'owasp/owasp_add.html'
        context = {
            'form': form
        }

    return render(request, template, context)

@login_required
def owasp_delete(request,pk):
    DB_OWASP.objects.filter(pk=pk).delete()
    return redirect('/owasp/list/')

#---------------------------------------------------
#                   CUSTOMERS
# --------------------------------------------------

@login_required
def customer_list(request):
    customers = Customer.objects.all()
    template = loader.get_template('customer/customer_list.html')
    context = {
        'customers':customers
    }
    return HttpResponse(template.render(context, request))

@login_required
def customer_add(request):
    if request.method == 'POST':
        form = AddCustomer(request.POST)
        
        if form.is_valid():
            customer = form.save(commit=False)            
            customer.save()
            
        return redirect('/customer/list')
        
    else:
        form = AddCustomer()
        template = 'customer/customer_add.html'
        context = {
            'form': form
        }

    return render(request, template, context)

@login_required
def customer_edit(request,pk):
    
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        form = AddCustomer(request.POST, instance=customer)
        
        if form.is_valid():
            customer = form.save(commit=False)            
            customer.save()
            
        return redirect('/customer/list')
    
    else:
        form = AddCustomer(instance=customer)
        template = 'customer/customer_add.html'
        context = {
            'form': form
        }

    return render(request, template, context)

@login_required
def customer_delete(request,pk):
    Customer.objects.filter(pk=pk).delete()
    next_url = request.GET.get('next', '/')
    return redirect(next_url)

@login_required
def customer_view(request,pk):

    Customer_query = get_object_or_404(Customer, pk=pk)
    Report_query = Report.objects.filter(customer=Customer_query).order_by('creation_date').reverse()
    count_customer_report = Report_query.count()
    customer_findings = {}
    count_customer_findings_total = 0
    count_customer_findings_critical_high = 0
    count_customer_findings_medium = 0

    for report in Report_query:
        Finding_query = Finding.objects.filter(report=report.id)
        count_customer_findings = Finding_query.count()
        customer_findings[report.id] = count_customer_findings
        count_customer_findings_total += count_customer_findings
        for finding in Finding_query:
            if finding.severity == 'High' or finding.severity == 'Critical':
                count_customer_findings_critical_high += 1
            elif finding.severity == 'Medium':
                count_customer_findings_medium += 1

    return render(request, 'customer/customer_view.html', {'pk': pk, 'Customer_query': Customer_query, 'Report_query': Report_query, 'count_customer_report': count_customer_report, 'customer_findings': count_customer_findings_total, 'count_customer_findings_critical_high': count_customer_findings_critical_high, 'count_customer_findings_medium': count_customer_findings_medium})

#---------------------------------------------------
#                   REPORT
# --------------------------------------------------
@login_required
def report_list(request):
    reports = Report.objects.all()
    template = loader.get_template('report/report_list.html')
    context = {
        'reports':reports
    }
    return HttpResponse(template.render(context, request))

@login_required
def report_add(request):
    today = datetime.date.today().strftime('%Y-%m-%d')
    report_id_format = "PEN-DOC" + str(datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S'))

    if request.method == 'POST':
        form = AddReport(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            audit_dates = request.POST['audit']
            split_audit_dates = audit_dates.split(" - ")
            report.audit_start = split_audit_dates[0]
            report.audit_end = split_audit_dates[1]
            report.save()           
            return redirect(reverse('report_view', kwargs={'pk' : report.pk}))
    else:
        form = AddReport()
        form.fields['report_id'].initial = report_id_format
        form.fields['report_date'].initial = today
        form.fields['audit'].initial = "%s - %s" % (today,today)
        
    return render(request, 'report/report_add.html', {
        'form': form
    })

@login_required
def report_edit(request,pk):

    report = get_object_or_404(Report, pk=pk)

    if request.method == 'POST':
        form = AddReport(request.POST, instance=report)
        if form.is_valid():
            report = form.save(commit=False)
            audit_dates = request.POST['audit']
            split_audit_dates = audit_dates.split(" - ")
            report.audit_start = split_audit_dates[0]
            report.audit_end = split_audit_dates[1]
            report.save()
            # CHANGE THIS WHEN VIEW IS DONE
            # return redirect(reverse('report_view', kwargs={'pk' : report.pk}))
            return redirect('/report/list')
    else:
        form = AddReport(instance=report)
        
    return render(request, 'report/report_add.html', {
        'form': form
    })

def generate_nmap_markdown(nmap_data):
    markdown_parts = []

    for host in nmap_data.get('hosts', []):
        ip_address = host.get('ip_address', 'Unknown IP Address')
        markdown_parts.append(f"## Host: {ip_address}\n\n")
        
        ports = host.get('ports', [])
        if not ports:
            markdown_parts.append("No open ports found.\n\n")
            continue
        
        for port in ports:
            port_id = port.get('port_id', 'unknown')
            state = port.get('state', 'unknown')
            service_name = port.get('service', {}).get('name', 'unknown')
            product = port.get('service', {}).get('product', 'unknown')
            
            markdown_parts.append(f"- **Port {port_id}** ({state}): {service_name} {product}\n")
            
            scripts = port.get('scripts', [])
            if scripts:
                for script in scripts:
                    script_id = script.get('id', 'unknown')
                    output = script.get('output', 'No output.')
                    markdown_parts.append(f"    - **{script_id}**: {output}\n")
            else:
                markdown_parts.append("    - No scripts found.\n")
        markdown_parts.append("\n")  # Add a newline for better separation

    return ''.join(markdown_parts)

@login_required
def report_view(request,pk):
    
    DB_report_query = get_object_or_404(Report, pk=pk)
    DB_finding_query = Finding.objects.filter(report=DB_report_query).order_by('-cvss_score')
    count_finding_query = DB_finding_query.count()
    
    DB_deliverable_query = Deliverable.objects.filter(report=pk).order_by('pk')

    if DB_report_query.nmap_scan != "":
        nmap_data = generate_nmap_markdown(json.loads(DB_report_query.nmap_scan))
    else:
        nmap_data = ""

    count_findings_critical = 0
    count_findings_high = 0
    count_findings_medium = 0
    count_findings_low = 0
    count_findings_info = 0
    count_findings_none = 0

    count_open_findings = 0
    count_closed_findings = 0

    cwe_rows = []
    owasp_rows = []
    
    for finding in DB_finding_query:
        # Only reporting Critical/High/Medium/Low/Info findings
        if finding.severity == 'None':
            count_findings_none += 1
        else:

            if finding.cwe:
                finding_cwe = f"CWE-{finding.cwe.cwe_id} - {finding.cwe.cwe_name}"
                cwe_rows.append(finding_cwe)

            if finding.owasp:
                finding_owasp = f"{finding.owasp.owasp_full_id} - {finding.owasp.owasp_name}"
                owasp_rows.append(finding_owasp)

            if finding.severity == 'Critical':
                count_findings_critical += 1
            elif finding.severity == 'High':
                count_findings_high += 1
            elif finding.severity == 'Medium':
                count_findings_medium += 1
            elif finding.severity == 'Low':
                count_findings_low += 1
            elif finding.severity == 'Info':
                count_findings_info += 1


            if finding.status == 'Open':
                count_open_findings += 1
            elif finding.status == 'Closed':
                count_closed_findings += 1


    cwe_cat = Counter(cwe_rows)
    cwe_categories = []

    for key_cwe, value_cwe in cwe_cat.items():
        fixed_key_cwe = '\n'.join(key_cwe[i:i+50] for i in range(0, len(key_cwe), 50))
        dict_cwe = {
            "value": value_cwe,
            "name": fixed_key_cwe
        }
        cwe_categories.append(dict_cwe)

    owasp_cat = Counter(owasp_rows)
    owasp_categories = []

    for key_owasp, value_owasp in owasp_cat.items():
        fixed_key_owasp = '\n'.join(key_owasp[i:i+55] for i in range(0, len(key_owasp), 55))
        dict_owasp = {
            "value": value_owasp,
            "name": fixed_key_owasp
        }
        owasp_categories.append(dict_owasp)


    return render(request, 'report/report_view.html', {'DB_report_query': DB_report_query, 'DB_finding_query': DB_finding_query, 'DB_deliverable_query': DB_deliverable_query, 'count_finding_query': count_finding_query, 'count_findings_critical': count_findings_critical, 'count_findings_high': count_findings_high, 'count_findings_medium': count_findings_medium, 'count_findings_low': count_findings_low, 'count_findings_info': count_findings_info, 'count_findings_none': count_findings_none, 'cwe_categories': cwe_categories, 'owasp_categories': owasp_categories, 'count_open_findings': count_open_findings, 'count_closed_findings': count_closed_findings, 'nmap_data': nmap_data})


@login_required
def report_delete(request,pk):
    Report.objects.filter(pk=pk).delete()
    next_url = request.GET.get('next', '/')
    return redirect(next_url)

@login_required
def report_finding(request,pk):
    DB_report_query = get_object_or_404(Report, pk=pk)
    DB_finding_query = Finding.objects.filter(report=DB_report_query).order_by('-cvss_score')
    success = request.session.pop('success', False)
    
    return render(request, 'report/report_finding.html', {
        'DB_report_query' : DB_report_query,
        'DB_finding_query' : DB_finding_query,
        'success' : success,
    })
    
@login_required
def uploadsummaryfindings(request, pk):
    DB_report_query = get_object_or_404(Report, pk=pk)

    if request.method == 'POST':

        # Severitybar
        summary_finding_file_base64 = request.POST['fileSeveritybar']
        format, summary_finding_file_str = summary_finding_file_base64.split(';base64,')
        summary_ext = format.split('/')[-1]
        dataimgSeveritybar = ContentFile(base64.b64decode(summary_finding_file_str))

         # OWASP Categories
        owasp_summary_categories_file_base64 = request.POST['file_owasp']
        formatf, owasp_summary_categories_finding_file_str = owasp_summary_categories_file_base64.split(';base64,')
        owasp_ext = formatf.split('/')[-1]
        dataOWASP = ContentFile(base64.b64decode(owasp_summary_categories_finding_file_str))

        DB_report_query.executive_summary_image = summary_finding_file_base64
        DB_report_query.owasp_categories_summary_image = owasp_summary_categories_file_base64
        DB_report_query.save()

        return HttpResponse('{"status":"success"}', content_type='application/json')
    else:
        return HttpResponseServerError('{"status":"fail"}', content_type='application/json')


@login_required
def reportdownloadpdf(request,pk):
    template_dir = os.path.join(settings.TEMPLATES_ROOT, 'pdf')
    template_pdf_dir = os.path.join(template_dir, 'default')
    
    DB_report_query = get_object_or_404(Report, pk=pk)
    DB_finding_query = Finding.objects.filter(report=DB_report_query).order_by('cvss_score').reverse()

    # ------------------------------------------------------------------------------------------------------------------------------------------------------    

    # Datetime
    now = datetime.datetime.utcnow()
    report_date = DB_report_query.report_date.strftime('%d-%m-%Y')

    # PDF filename
    file_name = DELIVERABLES_CONFIG['report_pdf_name'] + '_' + DB_report_query.title + '_' +  str(datetime.datetime.utcnow().strftime('%Y%m%d%H%M')).replace('/', '') + '.pdf'

    # INIT
    template_findings = template_appendix = pdf_finding_summary = ''
    md_author = DELIVERABLES_CONFIG['md_author']
    md_subject = DELIVERABLES_CONFIG['md_subject']
    md_website = DELIVERABLES_CONFIG['md_website']
    
    # Appendix
    for finding in DB_finding_query:
        if finding.severity == "None":
            continue
        if finding.appendix_finding.all():
            template_appendix = ('# Additional Notes') + "\n\n"
    
    # IMAGES
    report_executive_summary_image = DB_report_query.executive_summary_image
    report_owasp_categories_image = DB_report_query.owasp_categories_summary_image
    
    counter_finding = counter_finding_critical = counter_finding_high = counter_finding_medium = counter_finding_low = counter_finding_info = 0

    for finding in DB_finding_query:

        # Only reporting Critical/High/Medium/Low/Info findings
        if finding.severity == 'None':
            pass
        else:
            counter_finding += 1
            template_appendix_in_finding = ''

            if finding.severity == 'Critical':
                counter_finding_critical += 1
                icon_finding = 'important'
                severity_color = 'criticalcolor'
                severity_box = 'criticalbox'
            elif finding.severity == 'High':
                counter_finding_high += 1
                icon_finding = 'highnote'
                severity_color = 'highcolor'
                severity_box = 'highbox'
            elif finding.severity == 'Medium':
                counter_finding_medium += 1
                icon_finding = 'mediumnote'
                severity_color = 'mediumcolor'
                severity_box = 'mediumbox'
            elif finding.severity == 'Low':
                counter_finding_low += 1
                icon_finding = 'lownote'
                severity_color = 'lowcolor'
                severity_box = 'lowbox'
            else:
                counter_finding_info += 1
                icon_finding = 'debugnote'
                severity_color = 'debugcolor'
                severity_box = 'infobox'

            pdf_finding_summary += render_to_string(os.path.join(template_pdf_dir, 'pdf_finding_summary.md'),{'finding': finding,'counter_finding': counter_finding, 'severity_box': severity_box})
            severity_color_finding = "\\textcolor{" + f"{severity_color}" +"}{" + f"{finding.severity}" + "}"

            # appendix
            if finding.appendix_finding.all():
                
                template_appendix_in_finding = ('**Additional notes**') + "\n\n"

                for appendix_in_finding in finding.appendix_finding.all():

                    pdf_appendix = render_to_string(os.path.join(template_pdf_dir, 'pdf_appendix.md'),{'appendix_in_finding': appendix_in_finding})

                    template_appendix += ''.join(pdf_appendix + "  \n\n")
                    template_appendix_in_finding += ''.join((appendix_in_finding.title) + "\n")
                    

                template_appendix_in_finding += ''.join("\\pagebreak")

            else:
                template_appendix_in_finding += ''.join("\\pagebreak")
            
            # finding
            pdf_finding = render_to_string(os.path.join(template_pdf_dir, 'pdf_finding.md'), {'finding': finding, 'icon_finding': icon_finding, 'severity_color': severity_color, 'severity_color_finding': severity_color_finding, 'template_appendix_in_finding': template_appendix_in_finding })
            template_findings += ''.join(pdf_finding)

    if DB_report_query.nmap_scan != "":    
        nmap_data = generate_nmap_markdown(json.loads(DB_report_query.nmap_scan))
    else:
        nmap_data = ""
    pdf_markdown_report = render_to_string(os.path.join(template_pdf_dir, 'pdf_header.yaml'), {'DB_report_query': DB_report_query, 'md_author': md_author, 'report_date': report_date, 'md_subject': md_subject, 'md_website': md_website, 'report_pdf_language': 'en', 'titlepagecolor': DELIVERABLES_CONFIG['titlepage-color'], 'titlepagetextcolor': DELIVERABLES_CONFIG['titlepage-text-color'], 'titlerulecolor': DELIVERABLES_CONFIG['titlepage-rule-color'], 'titlepageruleheight': DELIVERABLES_CONFIG['titlepage-rule-height'] })
    pdf_markdown_report += render_to_string(os.path.join(template_pdf_dir, 'pdf_report.md'), {'DB_report_query': DB_report_query, 'template_findings': template_findings, 'report_executive_summary_image': report_executive_summary_image, 'report_owasp_categories_image': report_owasp_categories_image, 'pdf_finding_summary': pdf_finding_summary, 'nmap_data' : nmap_data, 'template_appendix': template_appendix})

    final_markdown = textwrap.dedent(pdf_markdown_report)

    header = render_to_string(os.path.join(template_pdf_dir, 'pdf_header.yaml'))
    final_markdown = header + final_markdown 
    
    pdf_file_output = os.path.join(settings.REPORTS_MEDIA_ROOT, 'pdf', file_name)
    
    PDF_HEADER_FILE = os.path.join(template_pdf_dir, 'pdf_header.tex')

    REPORTING_LATEX_FILE = os.path.join(template_pdf_dir,'report_default.tex')
    
    pypandoc.convert_text(final_markdown, to='pdf', outputfile=pdf_file_output, format='md',extra_args=[
                                        '-H', PDF_HEADER_FILE,
                                        '--filter', 'pandoc-latex-environment',
                                        '--from', 'markdown+yaml_metadata_block+raw_html',
                                        '--template', REPORTING_LATEX_FILE,
                                        '--pdf-engine', DELIVERABLES_CONFIG['pdf_engine'],])
    
    deliverable = Deliverable(report=DB_report_query, filename=file_name, generation_date=now, filetype='pdf')
    deliverable.save()
        
    with open(pdf_file_output, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_name)  # or 'attachment; filename=' 'inline; filename='
        return response

@login_required
def reportdownloadexcel(request, pk):
    template_dir = os.path.join(settings.TEMPLATES_ROOT, 'excel')
    
    now = datetime.datetime.utcnow()
    
    DB_report_query = get_object_or_404(Report, pk=pk)
    DB_finding_query = Finding.objects.filter(report=DB_report_query).order_by('cvss_score').reverse()
    
    file_name = DELIVERABLES_CONFIG['report_excel_name'] + '_' + DB_report_query.title + '_' +  str(datetime.datetime.utcnow().strftime('%Y%m%d%H%M')).replace('/', '') + '.xlsx'
    
    wb = openpyxl.load_workbook(os.path.join(template_dir, 'Finding_Remediation_Checklist.xlsx'))
    ws = wb["Checklist"]
    
    status_validation = DataValidation(type="list", formula1='"Open,In-Progress,Closed"', showDropDown=False)
    ws.add_data_validation(status_validation)
    
    severity_styles = {
        'Critical': PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid'),  
        'High': PatternFill(start_color='F20000', end_color='F20000', fill_type='solid'),  
        'Medium': PatternFill(start_color='FC7F03', end_color='FC7F03', fill_type='solid'),  
        'Low': PatternFill(start_color='05B04F', end_color='05B04F', fill_type='solid'),
        'Info': PatternFill(start_color='45A7F7', end_color='45A7F7', fill_type='solid'),
    }
    
    status_styles = {
        'Open': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),  
        'In-Progress': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),  
        'Closed': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),  
    }
    
    row_num = 4
    
    for finding in DB_finding_query:
        if finding.severity == "None":
            continue
        else:
            finding_cell = ws.cell(row=row_num, column=2, value=finding.title)
            severity_cell = ws.cell(row=row_num, column=3, value=finding.severity)
            status_cell = ws.cell(row=row_num, column=5, value=finding.status)
            
            status_validation.add(status_cell)
            
            if finding.severity in severity_styles:
                severity_cell.fill = severity_styles[finding.severity]
                
            ws.conditional_formatting.add(f"E{row_num}", CellIsRule(operator='equal', formula=['"Open"'], fill=status_styles["Open"]))
            ws.conditional_formatting.add(f"E{row_num}", CellIsRule(operator='equal', formula=['"In-Progress"'], fill=status_styles["In-Progress"]))
            ws.conditional_formatting.add(f"E{row_num}", CellIsRule(operator='equal', formula=['"Closed"'], fill=status_styles["Closed"]))
                
            status_validation.add(status_cell)
            
            row_num += 1
            
    excel_file_output = os.path.join(settings.REPORTS_MEDIA_ROOT, 'excel', file_name)
    
    deliverable = Deliverable(report=DB_report_query, filename=file_name, generation_date=now, filetype='excel')
    deliverable.save()
    
    wb.save(excel_file_output)
    
    with open(excel_file_output, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_name)
        
        return response
    
# ----------------------------------------------------------------------
#                           Deliverables
# ----------------------------------------------------------------------

@login_required
def deliverable_list(request):
    DB_deliverable_query = Deliverable.objects.order_by('pk').all()
    return render(request, 'deliverable/deliverable_list.html', {'DB_deliverable_query': DB_deliverable_query})

@login_required
def deliverable_download(request, pk):
    deliverable = get_object_or_404(Deliverable, pk=pk)
    file_path = os.path.join(settings.REPORTS_MEDIA_ROOT, deliverable.filetype, deliverable.filename )

    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            if deliverable.filetype == 'pdf':
                content_type="application/pdf"
            elif deliverable.filetype == 'excel':
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                content_type="application/octet-stream"
            response = HttpResponse(fh.read(), content_type=content_type)
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_path)
            return response

    raise Http404

@login_required
def deliverable_delete(request, pk):
    deliverable = get_object_or_404(Deliverable, pk=pk)
    file_path = os.path.join(settings.REPORTS_MEDIA_ROOT, deliverable.filetype, deliverable.filename )
    
    if os.path.exists(file_path):
        os.remove(file_path)
        next_url = request.GET.get('next', '/')
        deliverable.delete()
        return redirect(next_url)
    
    raise Http404    


#---------------------------------------------------
#                   NMAP
# --------------------------------------------------

def xml_to_dict_nmap(xml_string):
    # Load and parse the XML file
    root = ET.fromstring(xml_string)
    scan_results = {"hosts": []}

    for host in root.findall('host'):
        host_dict = {
            "ip_address": host.find('address').get('addr'),
            "ports": []
        }

        for port in host.find('ports').findall('port'):
            port_dict = {
                "port_id": port.get('portid'),
                "state": port.find('state').get('state'),
                "service": {
                    "name": port.find('service').get('name', 'unknown') if port.find('service') is not None else 'unknown',
                    "product": port.find('service').get('product', 'unknown') if port.find('service') is not None else 'unknown'
                },
                "scripts": []
            }

            for script in port.findall('script'):
                script_dict = {
                    "id": script.get('id'),
                    "output": script.get('output')
                }
                port_dict["scripts"].append(script_dict)

            host_dict["ports"].append(port_dict)

        scan_results["hosts"].append(host_dict)

    return scan_results


@login_required
def upload_and_parse_nmap(request,pk):
    report_query = get_object_or_404(Report, pk=pk)
    if request.method == 'POST':
        form = UploadNmapForm(request.POST, request.FILES)
        if form.is_valid():
            xml_file = request.FILES['nmap_file']
            xml_data = xml_file.read()
            parsed_xml = xml_to_dict_nmap(xml_data)
            json_data = json.dumps(parsed_xml)
            report_query.nmap_scan = json_data
            report_query.save()
            
            return redirect(reverse('report_view', kwargs={'pk' : pk}))  
    else:
        form = UploadNmapForm()
    return render(request, 'report/report_import_nmap.html', {'form': form, 'DB_report_query' : report_query})

@login_required
def remove_nmap_scan(request,pk):
    report_query = get_object_or_404(Report, pk=pk)
    report_query.nmap_scan = ""
    report_query.save()
    
    return redirect(reverse('report_view', kwargs={'pk' : pk}))


#---------------------------------------------------
#                   OPENVAS
# --------------------------------------------------

def parse_openvas_xml(xml_string):
    root = ET.fromstring(xml_string)

    findings = []

    for result in root.findall('.//result'):
        tags_text = result.find('.//tags').text if result.find('.//tags') is not None else ""
        tags_dict = dict(tag.split('=', 1) for tag in tags_text.split('|') if '=' in tag)
        result_id = result.attrib.get('id', "N/A")
        severity_text = result.find('severity').text if result.find('severity') is not None else "N/A"
        try:
            severity = float(severity_text)
        except ValueError:
            print(f"Could not convert severity '{severity_text}' to float.")
            severity = None

        finding = {
            'id': result_id,
            'name': result.find('.//name').text if result.find('.//name') is not None else "N/A",
            'description': tags_dict.get('summary', "N/A").strip() + "\n" + tags_dict.get('impact', "N/A").strip(),
            # 'impact': tags_dict.get('impact', "N/A").strip(),
            'solution': result.find('.//solution').text if result.find('.//solution') is not None else "N/A",
            'host': result.find('.//host').text if result.find('.//host') is not None else "N/A",
            'references': [ref.attrib['id'] for ref in result.findall('.//refs/ref')] if result.findall('.//refs/ref') is not None else [],
            'poc' : result.find('.//description').text if result.find('.//description') is not None else "N/A",
            'severity' : result.find('.//threat').text if result.find('.//threat') is not None else "N/A",
            'cvss_score' : severity
        }
        findings.append(finding)

    return findings

@login_required
def upload_and_parse_openvas(request,pk):
    report_query = get_object_or_404(Report, pk=pk)
    if request.method == 'POST':
        openvas_form = UploadOpenVASForm(request.POST, request.FILES)
        if openvas_form.is_valid():
            openvas_file = request.FILES['openvas_file']
            openvas_data = openvas_file.read()
            parsed_openvas = parse_openvas_xml(openvas_data)
            json_data = json.dumps(parsed_openvas, indent=4)
            report_query.openvas_scan = json_data
            report_query.save()
            
            return redirect(reverse('upload_and_parse_openvas', kwargs={'pk' : pk}))
        else:
            selected_ids = request.POST.getlist('vulnerability_ids')
            vulnerabilities = json.loads(report_query.openvas_scan)
            report_finding_query = Finding.objects.filter(report=report_query)
            queryset_names = report_finding_query.values_list('title', flat=True)
            for vulnerability in vulnerabilities:  # vulnerabilities list of dicts
                if str(vulnerability['id']) in selected_ids:
                    if vulnerability['name'] in queryset_names:
                        continue
                    else:
                        new_finding_title = vulnerability['name']
                        new_finding_description = vulnerability['description'].replace("\n", "")
                        new_finding_location = vulnerability['host']
                        # new_finding_impact = vulnerability['impact']
                        new_finding_recommendation = vulnerability['solution']
                        formatted_urls = ["- " + url + "\n" for url in vulnerability['references']]
                        new_finding_references = "".join(formatted_urls)
                        new_finding_poc = vulnerability['poc']
                        new_finding_severity = vulnerability['severity']
                        new_finding_cvss_score = float(vulnerability['cvss_score'])
                        new_finding_report = report_query
                        new_finding_finding_id = uuid.uuid4()
                        
                        new_finding = Finding(
                            finding_id = new_finding_finding_id,
                            title = new_finding_title,
                            description = new_finding_description,
                            location = new_finding_location,
                            # impact = new_finding_impact,
                            recommendation = new_finding_recommendation,
                            references = new_finding_references,
                            poc = new_finding_poc,
                            report = new_finding_report,
                            cvss_score = new_finding_cvss_score,
                            severity = new_finding_severity
                        )
                        
                        new_finding.save()
                else:
                    if vulnerability['name'] in queryset_names:
                        finding_to_delete = Finding.objects.filter(title=vulnerability['name'])
                        finding_to_delete.delete()
                    
            return redirect(reverse('report_view', kwargs={'pk' : pk}))
        
    else:
        openvas_form = UploadOpenVASForm()
        matching_vulnerabilities = []
        if report_query.openvas_scan != "":
            vulnerabilities = json.loads(report_query.openvas_scan)
            report_finding_query = Finding.objects.filter(report=report_query)
            queryset_names = report_finding_query.values_list('title', flat=True)
            for vulnerability in vulnerabilities:
                if vulnerability['name'] in queryset_names:
                    matching_vulnerabilities.append(vulnerability['name'])
        else:
            vulnerabilities = ""

    return render(request, 'report/report_import_openvas.html', 
                  {
                      'openvas_form': openvas_form,
                      'vulnerabilities' : vulnerabilities,
                      'DB_report_query' : report_query,
                      'matching_vulnerabilities' : matching_vulnerabilities
                      })
    
@login_required
def remove_openvas_scan(request,pk):
    report_query = get_object_or_404(Report, pk=pk)
    report_query.openvas_scan = ""
    report_query.save()
    
    return redirect(reverse('report_view', kwargs={'pk' : pk}))

#---------------------------------------------------
#                   APPENDIX
# --------------------------------------------------

@login_required
def reportappendix_list(request,pk):
    report_query = get_object_or_404(Report, pk=pk)
    finding_query = Finding.objects.filter(report=report_query).order_by('cvss_score').reverse()
    appendix_query = Appendix.objects.filter(finding__in=finding_query)

    count_appendix_query = appendix_query.count()

    return render(request, 'appendix/reportappendix_list.html', {'report_query': report_query, 'finding_query': finding_query, 'appendix_query': appendix_query, 'count_appendix_query': count_appendix_query})


@login_required
def appendix_add(request,pk):

    report_query = get_object_or_404(Report, pk=pk)

    if request.method == 'POST':
        form = AppendixForm(request.POST, reportpk=pk)
        if form.is_valid():
            appendix = form.save(commit=False)            
            finding_pk = form['finding'].value()
            appendix.save()
            appendix.finding.add(finding_pk)

            if '_finish' in request.POST:
                return redirect('reportappendix_list', pk=pk)
            elif '_next' in request.POST:
                return redirect('appendix_add', pk=pk)
    else:
        form = AppendixForm(reportpk=pk)
        form.fields['description'].initial = 'TBD'


    return render(request, 'appendix/appendix_add.html', {
        'form': form, 'report_query': report_query
    })
    
    
@login_required
def appendix_edit(request,pk):

    appendix = get_object_or_404(Appendix, pk=pk)
    finding_pk = appendix.finding.first().pk
    finding_query = get_object_or_404(Finding, pk=finding_pk)

    report = finding_query.report
    report_query = get_object_or_404(Report, pk=report.pk)

    if request.method == 'POST':
        form = AppendixForm(request.POST, instance=appendix, reportpk=report.pk)
        if form.is_valid():
            appendix = form.save(commit=False)
            new_finding_pk = form['finding'].value()
            New_finding = Finding.objects.filter(pk=new_finding_pk)
            appendix.save()
            appendix.finding.set(New_finding, clear=True)

            if '_finish' in request.POST:
                return redirect('reportappendix_list', pk=report.pk)
            elif '_next' in request.POST:
                return redirect('appendix_add', pk=report.pk)
    else:
        form = AppendixForm(reportpk=report.pk, instance=appendix, initial={'finding': finding_pk})

    return render(request, 'appendix/appendix_add.html', {
        'form': form, 'report_query': report_query
    })
    
    
@login_required
def appendix_duplicate(request,pk,appendixpk):

    report_query = get_object_or_404(Report, pk=pk)
    appendix_query = get_object_or_404(Appendix, pk=appendixpk)

    if request.method == 'POST':
        form = AppendixForm(request.POST, reportpk=pk)
        if form.is_valid():
            appendix = form.save(commit=False)            
            finding_pk = form['finding'].value()
            appendix.save()
            appendix.finding.add(finding_pk)

            if '_finish' in request.POST:
                return redirect('reportappendix_list', pk=pk)
            elif '_next' in request.POST:
                return redirect('appendix_add', pk=pk)
    else:
        form = AppendixForm(reportpk=pk)
        form.fields['description'].initial = appendix_query.description
        form.fields['title'].initial = appendix_query.title


    return render(request, 'appendix/appendix_add.html', {
        'form': form, 'report_query': report_query
    })

    
@login_required
def appendix_delete(request,pk):

    Appendix.objects.filter(pk=pk).delete()
    next_url = request.GET.get('next', '/')
    return redirect(next_url)


@login_required
def appendix_view(request,pk):
    appendix = get_object_or_404(Appendix, pk=pk)
    finding_pk = appendix.finding.first().pk
    finding_query = get_object_or_404(Finding, pk=finding_pk)

    return render(request, 'appendix/appendix_view.html', {'finding_query': finding_query, 'appendix_query': appendix})


# ------------------------------------------------
#                     TEST
# ------------------------------------------------

